import csv
import os
from typing import List, Dict, Any, Optional, Tuple


_ALIASES = {
    'dap': 'DAP',
    'dap_': 'DAP',
    'doseairproduct': 'DAP',
    'dapagycm2': 'DAP_(A)_Gy_cm2',
    'dapatotalgycm2': 'DAP_Gy_cm2',
    'dapbgycm2': 'DAP_(B)_Gy_cm2',
    'reference_point_dose': 'Reference_Point_Dose',
    'referencepointdose': 'Reference_Point_Dose',
    'referencepointdosemgy': 'Reference_Point_Dose',
    'referencepointdosetotalmgy': 'Reference_Point_Dose',
    'referencepointdoseamgy': 'Reference_Point_Dose_(A)_mGy',
    'referencepointdosebmgy': 'Reference_Point_Dose_(B)_mGy',
    'rpd': 'Reference_Point_Dose',
    'kvp': 'kVp',
    'kvp_': 'kVp',
    'tubevoltage': 'kVp',
    'source_to_detector_distance': 'Source_To_Detector_Distance',
    'sourcetodetectordistance': 'Source_To_Detector_Distance',
    'sourcetodetectordistancerfmm': 'Source_To_Detector_Distance',
    'sourcetodetectordistancerf': 'Source_To_Detector_Distance',
    'sid': 'Source_To_Detector_Distance',
    'source_to_isocenter_distance': 'Source_To_Isocenter_Distance',
    'sourcetoisocenterdistance': 'Source_To_Isocenter_Distance',
    'sourcetoisocenterdistancerfmm': 'Source_To_Isocenter_Distance',
    'sourcetoisocenterdistancerf': 'Source_To_Isocenter_Distance',
    'sad': 'Source_To_Isocenter_Distance',
    'collimated_field_area': 'Collimated_Field_Area',
    'collimatedfieldarea': 'Collimated_Field_Area',
    'collimatedfieldarearfcm2': 'Collimated_Field_Area',
    'collimatedfieldarearfcm': 'Collimated_Field_Area',
    'collimatedfieldaracm2': 'Collimated_Field_Area',
    'collimatedfieldaracm': 'Collimated_Field_Area',
    'fieldarea': 'Collimated_Field_Area',
    'collimatedarea': 'Collimated_Field_Area',
    'table_lateral_position': 'Table_Lateral_Position',
    'tablelateralposition': 'Table_Lateral_Position',
    'tablelateralpositionmm': 'Table_Lateral_Position',
    'tablelateral': 'Table_Lateral_Position',
    'table_longitudinal_position': 'Table_Longitudinal_Position',
    'tablelongitudinalposition': 'Table_Longitudinal_Position',
    'tablelongitudinalpositionmm': 'Table_Longitudinal_Position',
    'tablelongitudinal': 'Table_Longitudinal_Position',
    'primary_angle': 'Primary_Angle_(RF)',
    'primaryangle': 'Primary_Angle_(RF)',
    'primaryanglerf': 'Primary_Angle_(RF)',
    'primary_angle_(rf)': 'Primary_Angle_(RF)',
    'secondary_angle': 'Secondary_Angle',
    'secondaryangle': 'Secondary_Angle',
    'secondaryanglerf': 'Secondary_Angle',
    'equipment': 'Equipment',
    # Keep Device separate from Equipment. In the example RDSRs, 'Equipment'
    # carries the room/unit name used for calibration lookup, while 'Device'
    # is a generic system label (e.g., AXIOM-Artis).
    'device': 'Device',
    'machine': 'Equipment',
    'manufacturer': 'Manufacturer',
    'make': 'Manufacturer',
    'oem': 'Manufacturer',
    'table_height_position': 'Table_Height_Position',
    'tableheightposition': 'Table_Height_Position',
    'tableheightpositionmm': 'Table_Height_Position',
    'tableheight': 'Table_Height_Position',
    'a#': 'A#',
    'accessionnumber': 'A#',
}


def _normalize_header(s: str) -> str:
    # Convert superscript digits to regular digits (e.g., ² -> 2)
    superscripts = str.maketrans('⁰¹²³⁴⁵⁶⁷⁸⁹', '0123456789')
    s = s.translate(superscripts)
    # Strip common encoding artifacts and degree symbols before normalization.
    s = s.replace('Â', '').replace('°', '').replace('º', '')
    try:
        import unicodedata

        s = unicodedata.normalize('NFKD', s)
        s = s.encode('ascii', 'ignore').decode('ascii')
    except Exception:
        pass
    return ''.join(ch for ch in s.lower() if ch.isalnum())


def _map_headers(headers: List[str]) -> Dict[int, str]:
    mapped: Dict[int, str] = {}
    used_names: set[str] = set()
    dup_counts: Dict[str, int] = {}
    for i, h in enumerate(headers):
        key = _normalize_header(h)
        proposed = _ALIASES.get(key, h.strip())
        # Avoid collisions where multiple source columns map to the same
        # normalized name.
        #
        # Important: keep the canonical prefix in the name so downstream
        # code can aggregate multiple columns (MATLAB-style) by prefix.
        # Example: multiple Reference_Point_Dose* columns should become
        # Reference_Point_Dose, Reference_Point_Dose__2, ...
        if proposed in used_names:
            dup_counts[proposed] = dup_counts.get(proposed, 1) + 1
            proposed = f"{proposed}__{dup_counts[proposed]}"
        else:
            dup_counts.setdefault(proposed, 1)
        mapped[i] = proposed
        used_names.add(proposed)
    return mapped


def list_excel_sheets(path: str) -> List[str]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        return list(wb.sheetnames)
    except Exception as exc:
        raise RuntimeError("Excel parsing requires openpyxl") from exc


def load_rdsr(path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """Load an RDSR-like CSV/XLSX and return structured acquisitions.

    Returns dict with keys: 'patient_dims' (or None), 'acquisitions' (list of dicts)
    Caller should prompt for patient dims if None.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    lower = path.lower()
    rows: List[List[Any]] = []
    sheet_name_used: Optional[str] = None
    available_sheets: List[str] = []
    if lower.endswith('.csv'):
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            for r in reader:
                rows.append(r)
    elif lower.endswith('.xls') or lower.endswith('.xlsx'):
        # Prefer openpyxl directly (streaming read) for speed and lower memory.
        # Fall back to pandas if openpyxl isn't available.
        try:
            import openpyxl

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            available_sheets = list(wb.sheetnames)
            if sheet_name is None:
                ws = wb.active
            else:
                if sheet_name not in wb.sheetnames:
                    raise KeyError(f"Sheet '{sheet_name}' not found; available: {wb.sheetnames}")
                ws = wb[sheet_name]
            sheet_name_used = ws.title

            for r in ws.iter_rows(values_only=True):
                rows.append(list(r))
        except Exception:
            try:
                import pandas as pd

                df = pd.read_excel(path, sheet_name=sheet_name, header=None)
                if isinstance(sheet_name, str):
                    sheet_name_used = sheet_name
                for r in df.values.tolist():
                    rows.append(list(r))
            except Exception as e:
                raise RuntimeError('Excel parsing requires openpyxl or pandas; install one or provide CSV') from e
    else:
        raise RuntimeError('Unsupported file type: ' + path)

    # Heuristic: find header row by locating a row with many alphabetic entries
    header_row = None
    for idx, r in enumerate(rows[:10]):
        joined = ' '.join(
            str(cell)
            for cell in r
            if cell is not None and (not isinstance(cell, str) or cell.strip() != '')
        )
        if any(k.lower() in joined.lower() for k in ['reference', 'dap', 'kvp', 'primary', 'secondary', 'table']):
            header_row = idx
            break
    if header_row is None:
        header_row = 0

    headers = [str(x).strip() if x is not None else '' for x in rows[header_row]]
    mapping = _map_headers(headers)
    column_map = []
    for idx, header in enumerate(headers):
        column_map.append(
            {
                "column_index": int(idx + 1),
                "source_header": header,
                "mapped_header": mapping.get(idx, ""),
            }
        )

    acquisitions = []
    def _cell_is_nonempty(cell: Any) -> bool:
        if cell is None:
            return False
        if isinstance(cell, str):
            return bool(cell.strip())
        return True

    def _to_float_or_value(val: Any) -> Any:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, str):
            s = val.strip()
            if s == '':
                return None
            try:
                return float(s)
            except Exception:
                return val
        return val

    for r in rows[header_row + 1:]:
        if not any(_cell_is_nonempty(cell) for cell in r):
            continue
        entry: Dict[str, Any] = {}
        for i, cell in enumerate(r):
            key = mapping.get(i)
            if not key:
                continue
            entry[key] = _to_float_or_value(cell)
        acquisitions.append(entry)

    return {
        'patient_dims': None,
        'acquisitions': acquisitions,
        'column_map': column_map,
        'header_row_index': int(header_row + 1),
        'sheet_name_used': sheet_name_used,
        'available_sheets': available_sheets,
    }


def extract_acquisitions(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Extract acquisition list from parsed RDSR data.
    
    Args:
        data: Dict returned by load_rdsr() with 'acquisitions' key.
        
    Returns:
        List of acquisition dicts, each with normalized keys for dose, geometry, and equipment.
    """
    return data.get('acquisitions', [])


def _normalize_manufacturer(value: Any) -> Tuple[str, str]:
    if value is None:
        display = "(unknown)"
    else:
        display = str(value).strip()
        if not display:
            display = "(unknown)"
    key = display.lower()
    return key, display


def get_manufacturer_summary(acquisitions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    counts: Dict[str, int] = {}
    display_by_key: Dict[str, str] = {}
    for acq in acquisitions:
        if not isinstance(acq, dict):
            continue
        mfr = acq.get('Manufacturer', acq.get('manufacturer'))
        key, display = _normalize_manufacturer(mfr)
        counts[key] = counts.get(key, 0) + 1
        display_by_key.setdefault(key, display)
    summary = [
        {
            'key': key,
            'display': display_by_key[key],
            'count': counts[key],
        }
        for key in counts
    ]
    summary.sort(key=lambda item: item['display'].lower())
    return summary


def _should_swap_by_default(key: str) -> bool:
    """Determine if a manufacturer should have lat/long swap enabled by default.
    
    GE and variants: swap recommended (True)
    Siemens and variants: swap not recommended (False)
    Others: swap not recommended (False)
    
    Args:
        key: Normalized (lowercase) manufacturer key.
        
    Returns:
        True if swap should be enabled by default for this manufacturer.
    """
    ge_variants = {'ge', 'general electric', 'ge medical systems', 'gems'}
    if key in ge_variants:
        return True
    return False


def apply_table_lat_long_swap(
    acquisitions: List[Dict[str, Any]],
    swap_by_mfr: Optional[Dict[str, bool]],
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    stats = {
        'swapped_rows': 0,
        'skipped_rows_missing_fields': 0,
    }
    if not swap_by_mfr:
        return acquisitions, stats

    swap_map = {str(k).lower(): bool(v) for k, v in swap_by_mfr.items()}

    for acq in acquisitions:
        if not isinstance(acq, dict):
            continue
        mfr = acq.get('Manufacturer', acq.get('manufacturer'))
        key, _ = _normalize_manufacturer(mfr)
        if not swap_map.get(key, False):
            continue
        if 'Table_Lateral_Position' in acq and 'Table_Longitudinal_Position' in acq:
            acq['Table_Lateral_Position'], acq['Table_Longitudinal_Position'] = (
                acq['Table_Longitudinal_Position'],
                acq['Table_Lateral_Position'],
            )
            stats['swapped_rows'] += 1
        else:
            stats['skipped_rows_missing_fields'] += 1
    return acquisitions, stats


def build_table_lat_long_swap_report(
    summary: List[Dict[str, Any]],
    swap_by_mfr: Optional[Dict[str, bool]],
    stats: Optional[Dict[str, int]] = None,
) -> Dict[str, Any]:
    selected = []
    counts: Dict[str, int] = {}
    swap_map: Dict[str, bool] = {}
    if swap_by_mfr is None:
        swap_by_mfr = {}

    for item in summary:
        display = str(item.get('display', '')).strip()
        key = str(item.get('key', '')).strip().lower()
        count = int(item.get('count', 0))
        if display:
            counts[display] = count
            selected_flag = bool(swap_by_mfr.get(key, False))
            swap_map[display] = selected_flag
            if selected_flag:
                selected.append(display)

    report = {
        'selected_manufacturers': selected,
        'manufacturer_counts': counts,
        'swap_by_manufacturer': swap_map,
        'selected_keys': sorted(k for k, v in swap_by_mfr.items() if v),
        'swapped_rows': int(stats.get('swapped_rows', 0)) if stats else 0,
        'skipped_rows_missing_fields': int(stats.get('skipped_rows_missing_fields', 0)) if stats else 0,
    }
    return report


def prompt_patient_dimensions() -> Tuple[float, float, float]:
    """Prompt user at runtime for patient AP, LAT, LEN dimensions.
    
    Returns:
        Tuple of (AP, LAT, LEN) in cm.
        
    Raises:
        ValueError: If input cannot be parsed as three floats.
    """
    default_ap, default_lat, default_len = 40.0, 30.0, 20.0
    prompt_text = (
        f"Enter patient dimensions (AP, LAT, LEN) in cm, comma-separated\n"
        f"or press Enter for defaults ({default_ap}, {default_lat}, {default_len}): "
    )
    try:
        resp = input(prompt_text).strip()
        if not resp:
            return default_ap, default_lat, default_len
        parts = [float(p.strip()) for p in resp.split(',')]
        if len(parts) != 3:
            print(f"Expected 3 values, got {len(parts)}; using defaults.")
            return default_ap, default_lat, default_len
        if any(p <= 0 for p in parts):
            print("Dimensions must be positive; using defaults.")
            return default_ap, default_lat, default_len
        return tuple(parts)
    except ValueError as e:
        print(f"Parse error: {e}; using defaults.")
        return default_ap, default_lat, default_len


def load_with_dimensions(
    path: str,
    *,
    sheet_name: Optional[str] = None,
    patient_dims: Optional[Tuple[float, float, float]] = None,
) -> Dict[str, Any]:
    """Load RDSR file and ensure patient dimensions are available.

    This is the convenience entry point used by regression tests.

    Args:
        path: Path to CSV/XLSX file.
        sheet_name: Optional Excel sheet name to read (XLS/XLSX only).
        patient_dims: Optional (AP, LAT, LEN) tuple; if None, prompts user.

    Returns:
        Dict with keys:
        - 'patient_dims': (AP, LAT, LEN)
        - 'acquisitions': list of acquisition dicts
    """
    data = load_rdsr(path, sheet_name=sheet_name)
    acquisitions = extract_acquisitions(data)

    if patient_dims is None:
        patient_dims = prompt_patient_dimensions()

    return {'patient_dims': patient_dims, 'acquisitions': acquisitions}
