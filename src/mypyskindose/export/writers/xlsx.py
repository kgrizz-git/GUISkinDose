"""XLSX report writer (§2). Clinical audit workbook via ``openpyxl``.

``render_xlsx_bytes(payload)`` for in-memory downloads; ``write_xlsx(payload,
path)`` for filesystem writes.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import cast

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

from mypyskindose.safe_output import atomic_write_private
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .._format import (
    COLOR_ERROR,
    COLOR_WARNING,
    CORRECTION_HEADER,
    OFFSET_LABELS,
    collect_alert_lines,
    correction_row,
    dosimetric_rows,
)
from mypyskindose.spreadsheet_safety import neutralize_spreadsheet_value

from ..models import ExamSection, ExportPayload

_BOLD = Font(bold=True)
_TITLE = Font(bold=True, size=16)
_AMBER = PatternFill("solid", fgColor=COLOR_WARNING)
_RED = PatternFill("solid", fgColor=COLOR_ERROR)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _new_sheet(wb: Workbook, title: str) -> Worksheet:
    return cast(Worksheet, wb.create_sheet(title))


def _autofit(ws: Worksheet) -> None:
    """Widen columns to their longest cell (prevents ### clipping)."""
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            longest = max(len(line) for line in str(cell.value).splitlines() or [""])
            widths[cell.column] = max(widths.get(cell.column, 0), longest)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 80)


def _write_rows(ws: Worksheet, rows, start_row: int = 1, *, header: bool = False) -> int:
    r = start_row
    for i, row in enumerate(rows):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=neutralize_spreadsheet_value(value))
            if header and i == 0:
                cell.font = _BOLD
        r += 1
    return r


def _overview_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = cast(Worksheet, wb.active)
    ws.title = "Overview"
    ws.sheet_view.showGridLines = True
    ws["A1"] = neutralize_spreadsheet_value(payload.meta.report_title)
    ws["A1"].font = _TITLE
    meta_rows = [
        ["Application", payload.meta.app_name],
        ["Package version", payload.meta.package_version],
        ["Report schema version", payload.meta.schema_version],
        ["Generated", payload.meta.generated_at.isoformat(timespec="seconds")],
        ["Execution context", payload.meta.execution_context],
        ["Exams", payload.provenance.exam_count],
        ["Source type", payload.provenance.source_type],
    ]
    r = _write_rows(ws, meta_rows, start_row=3)

    r += 1
    ws.cell(row=r, column=1, value="Executive alerts").font = _BOLD
    r += 1
    for text, severity in collect_alert_lines(payload):
        cell = ws.cell(row=r, column=1, value=neutralize_spreadsheet_value(text))
        cell.alignment = _WRAP
        if severity == "error":
            cell.fill = _RED
        elif severity == "warning":
            cell.fill = _AMBER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Cumulative summary").font = _BOLD
    r += 1
    r = _write_rows(ws, dosimetric_rows(payload.cumulative.metrics), start_row=r)
    _autofit(ws)


def _results_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = _new_sheet(wb, "Results")
    ws.sheet_view.showGridLines = True
    if payload.is_multi_exam:
        headers = ["Metric"] + [e.exam_id for e in payload.exams] + ["Cumulative"]
        metric_names = [row[0] for row in dosimetric_rows(payload.cumulative.metrics)]
        per_exam_cols = [{row[0]: row[1] for row in dosimetric_rows(e.metrics)} for e in payload.exams]
        cum_col = {row[0]: row[1] for row in dosimetric_rows(payload.cumulative.metrics)}
        rows: list[list[str]] = [headers]
        for name in metric_names:
            row_vals = [str(col.get(name, "N/A")) for col in per_exam_cols]
            rows.append([name, *row_vals, str(cum_col.get(name, "N/A"))])
        _write_rows(ws, rows, header=True)
    else:
        rows = [["Metric", "Value"]] + dosimetric_rows(payload.cumulative.metrics)
        _write_rows(ws, rows, header=True)
    _autofit(ws)


def _settings_block(exam: ExamSection) -> list[list[str]]:
    rows: list[list[str]] = [["Setting", "Value"]]
    for key, value in exam.settings.items():
        if key in ("phantom", "patient_offset"):
            continue
        rows.append([key, str(value)])
    ph = exam.settings.get("phantom", {})
    for key, value in ph.items():
        rows.append([f"phantom.{key}", str(value)])
    off = exam.settings.get("patient_offset", {})
    for key, value in off.items():
        rows.append([str(OFFSET_LABELS.get(key, key)), str(value)])
    # Equipment + coordinate corrections
    rows.append(["Manufacturer", exam.manufacturer or "N/A"])
    rows.append(["Model", exam.model or "N/A"])
    rows.append(["Normalization profile", exam.normalization_profile or "N/A"])
    toggles = exam.coordinate.get("toggles", {})
    for key, value in toggles.items():
        rows.append([f"coord.{key}", str(value)])
    return rows


def _settings_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = _new_sheet(wb, "Equipment & Settings" if payload.is_multi_exam else "Settings")
    ws.sheet_view.showGridLines = True
    r = 1
    for exam in payload.exams:
        if payload.is_multi_exam:
            ws.cell(row=r, column=1, value=neutralize_spreadsheet_value(f"--- Exam {exam.exam_id} ---")).font = _BOLD
            r += 1
        r = _write_rows(ws, _settings_block(exam), start_row=r, header=True)
        r += 1
    _autofit(ws)


def _corrections_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = _new_sheet(wb, "Corrections")
    ws.sheet_view.showGridLines = True
    r = 1
    if payload.is_multi_exam:
        for exam in payload.exams:
            ws.cell(row=r, column=1, value=neutralize_spreadsheet_value(f"--- Exam {exam.exam_id} ---")).font = _BOLD
            r += 1
            rows = [CORRECTION_HEADER] + [correction_row(s) for s in exam.corrections]
            r = _write_rows(ws, rows, start_row=r, header=True)
            r += 1
        ws.cell(row=r, column=1, value=neutralize_spreadsheet_value("--- Cumulative (kerma-weighted) ---")).font = _BOLD
        r += 1
    rows = [CORRECTION_HEADER] + [correction_row(s) for s in payload.cumulative.corrections]
    _write_rows(ws, rows, start_row=r, header=True)
    _autofit(ws)


def _warnings_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = _new_sheet(wb, "Warnings")
    ws.sheet_view.showGridLines = True
    r = 1
    for text, severity in collect_alert_lines(payload):
        cell = ws.cell(row=r, column=1, value=neutralize_spreadsheet_value(text))
        cell.alignment = _WRAP
        if severity == "error":
            cell.fill = _RED
        elif severity == "warning":
            cell.fill = _AMBER
        r += 1
    for exam in payload.exams:
        for msg in exam.warnings:
            ws.cell(row=r, column=1, value=neutralize_spreadsheet_value(f"[{exam.exam_id}] {msg}")).alignment = _WRAP
            r += 1
    _autofit(ws)


def _images_sheet(wb: Workbook, payload: ExportPayload) -> None:
    ws = _new_sheet(wb, "Images")
    ws.sheet_view.showGridLines = True
    row = 1
    for entry in payload.images:
        ws.cell(row=row, column=1, value=neutralize_spreadsheet_value(entry.label)).font = _BOLD
        row += 1
        if entry.png_bytes is not None:
            img = XLImage(io.BytesIO(entry.png_bytes))
            # Scale down large renders to fit without obscuring tables.
            if img.width > 720:
                ratio = 720 / img.width
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
            ws.add_image(img, f"A{row}")
            row += max(int(img.height / 18) + 2, 12)
        else:
            ws.cell(row=row, column=1, value=neutralize_spreadsheet_value(entry.error_message or "Image unavailable"))
            row += 2
    _autofit(ws)


def build_workbook(payload: ExportPayload) -> Workbook:
    wb = Workbook()
    _overview_sheet(wb, payload)
    _results_sheet(wb, payload)
    _settings_sheet(wb, payload)
    _corrections_sheet(wb, payload)
    _warnings_sheet(wb, payload)
    _images_sheet(wb, payload)
    return wb


def render_xlsx_bytes(payload: ExportPayload) -> bytes:
    buf = io.BytesIO()
    build_workbook(payload).save(buf)
    return buf.getvalue()


def write_xlsx(payload: ExportPayload, path: Path) -> None:
    atomic_write_private(path, render_xlsx_bytes(payload))
