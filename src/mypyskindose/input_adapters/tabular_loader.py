"""Low-level file readers for CSV, TSV, and XLSX.

All functions return the raw DataFrame with header=None (every row included,
all values as str/object). Callers are responsible for header detection,
column mapping, and type coercion.

Each function also returns encoding/delimiter metadata needed for provenance.

Excel workbooks (``.xlsx`` / ``.xlsm``) are ZIP containers. Before materializing
a sheet, readers enforce uncompressed-member budgets and a row×column/cell
budget so a crafted workbook cannot bypass the compressed upload size cap
(CWE-409 / zip bomb).
"""

from __future__ import annotations

import csv
import zipfile
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")

# Post-decompress budgets for OOXML workbooks (compressed uploads may still be
# within the GUI 64 MiB cap while inflating far beyond process memory).
MAX_XLSX_UNCOMPRESSED_MEMBER_BYTES = 32 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_TOTAL_BYTES = 64 * 1024 * 1024
MAX_XLSX_ROWS = 100_000
MAX_XLSX_COLS = 512
MAX_XLSX_CELLS = 2_000_000


@dataclass
class _RawLoad:
    raw_df: pd.DataFrame
    encoding: str
    delimiter: str | None  # None for xlsx


def _sniff_delimiter(path: Path, encoding: str) -> str:
    with path.open(encoding=encoding, newline="", errors="replace") as f:
        sample = f.read(8192)
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _try_read_csv(path: Path, delimiter: str, encoding: str) -> pd.DataFrame:
    return pd.read_csv(
        path,
        sep=delimiter,
        header=None,
        dtype=str,
        encoding=encoding,
        keep_default_na=False,
        skip_blank_lines=False,
    )


def read_csv(path: Path | str) -> _RawLoad:
    """Read a CSV file with encoding fallback and delimiter sniffing."""
    path = Path(path)
    last_exc: Exception = RuntimeError("No encodings tried")

    for enc in _ENCODINGS:
        try:
            delimiter = _sniff_delimiter(path, enc)
            raw_df = _try_read_csv(path, delimiter, enc)
            return _RawLoad(raw_df=raw_df, encoding=enc, delimiter=delimiter)
        except UnicodeDecodeError as exc:
            last_exc = exc
        except Exception as exc:
            last_exc = exc
            break

    raise ValueError("Could not read CSV input with any supported encoding.") from last_exc


def read_tsv(path: Path | str) -> _RawLoad:
    """Read a TSV file with encoding fallback (tab delimiter assumed)."""
    path = Path(path)
    last_exc: Exception = RuntimeError("No encodings tried")

    for enc in _ENCODINGS:
        try:
            raw_df = _try_read_csv(path, "\t", enc)
            return _RawLoad(raw_df=raw_df, encoding=enc, delimiter="\t")
        except UnicodeDecodeError as exc:
            last_exc = exc
        except Exception as exc:
            last_exc = exc
            break

    raise ValueError("Could not read TSV input with any supported encoding.") from last_exc


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value)


_ZIP_READ_CHUNK_BYTES = 64 * 1024


def _raise_xlsx_size_exceeded(*, member: bool) -> None:
    if member:
        raise ValueError("Excel workbook member exceeds the maximum allowed uncompressed size.")
    raise ValueError("Excel workbook exceeds the maximum allowed uncompressed size.")


def _count_decompressed_member(archive: zipfile.ZipFile, info: zipfile.ZipInfo, *, total_so_far: int) -> int:
    """Inflate one ZIP member and return its byte count, aborting over budget."""
    # Cheap reject on declared size before touching the payload.
    if info.file_size > MAX_XLSX_UNCOMPRESSED_MEMBER_BYTES:
        _raise_xlsx_size_exceeded(member=True)
    if total_so_far + max(info.file_size, 0) > MAX_XLSX_UNCOMPRESSED_TOTAL_BYTES:
        _raise_xlsx_size_exceeded(member=False)

    member_bytes = 0
    with archive.open(info, "r") as member:
        while True:
            chunk = member.read(_ZIP_READ_CHUNK_BYTES)
            if not chunk:
                break
            member_bytes += len(chunk)
            if member_bytes > MAX_XLSX_UNCOMPRESSED_MEMBER_BYTES:
                _raise_xlsx_size_exceeded(member=True)
            if total_so_far + member_bytes > MAX_XLSX_UNCOMPRESSED_TOTAL_BYTES:
                _raise_xlsx_size_exceeded(member=False)
    return member_bytes


def assert_xlsx_zip_within_budget(path: Path) -> None:
    """Reject OOXML workbooks whose decompressed members exceed size budgets.

    Declared central-directory sizes are checked first, then each member is
    inflated through ``ZipFile.open`` and counted so a forged ``file_size``
    cannot bypass the cap when openpyxl later decompresses the same archive.
    """
    try:
        with zipfile.ZipFile(path) as archive:
            total_uncompressed = 0
            for info in archive.infolist():
                if info.is_dir():
                    continue
                total_uncompressed += _count_decompressed_member(
                    archive, info, total_so_far=total_uncompressed
                )
    except ValueError:
        raise
    except (zipfile.BadZipFile, zipfile.LargeZipFile, OSError, RuntimeError) as exc:
        raise ValueError("Invalid Excel workbook.") from exc


def _select_worksheet(workbook, sheet_name: str | int):
    if isinstance(sheet_name, int):
        try:
            return workbook.worksheets[sheet_name]
        except IndexError as exc:
            raise ValueError("Excel sheet index is out of range.") from exc
    try:
        return workbook[sheet_name]
    except KeyError as exc:
        raise ValueError("Excel sheet name was not found.") from exc


def _read_excel_rows(path: Path, sheet_name: str | int) -> list[list[str]]:
    """Stream one sheet in read-only mode and enforce row/column/cell budgets.

    Uses ``data_only=False`` deliberately: tabular DoseTrack/Radimetrics-style
    exports store literal values, and this matches the prior ``pd.read_excel``
    openpyxl default. ``data_only=True`` returns ``None`` for formulas without a
    cached calculated value and would drop those cells.
    """
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        worksheet = _select_worksheet(workbook, sheet_name)
        rows: list[list[str]] = []
        cell_count = 0
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if row_idx > MAX_XLSX_ROWS:
                raise ValueError("Excel sheet exceeds the maximum allowed number of rows.")
            values = list(row)
            if len(values) > MAX_XLSX_COLS:
                raise ValueError("Excel sheet exceeds the maximum allowed number of columns.")
            cell_count += len(values)
            if cell_count > MAX_XLSX_CELLS:
                raise ValueError("Excel sheet exceeds the maximum allowed cell budget.")
            rows.append([_cell_to_str(value) for value in values])
        return rows
    finally:
        workbook.close()


def read_excel(path: Path | str, sheet_name: str | int = 0) -> _RawLoad:
    """Read an XLSX/XLSM file with header=None so all rows are returned.

    Enforces uncompressed ZIP member budgets and a streamed row×column/cell
    budget before building the full DataFrame.
    """
    path = Path(path)
    assert_xlsx_zip_within_budget(path)
    rows = _read_excel_rows(path, sheet_name)
    if not rows:
        raw_df = pd.DataFrame()
    else:
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        raw_df = pd.DataFrame(normalized)
    return _RawLoad(raw_df=raw_df, encoding="utf-8", delimiter=None)


def load(
    path: Path | str,
    sheet_name: str | int = 0,
) -> _RawLoad:
    """Dispatch to the correct reader based on file suffix."""
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".tsv":
        return read_tsv(path)
    if suffix in (".xlsx", ".xlsm"):
        return read_excel(path, sheet_name=sheet_name)
    raise ValueError(
        f"Unsupported file suffix {suffix!r}. "
        "Tabular adapter handles .csv, .tsv, .xlsx, .xlsm."
    )
