"""Phase 5 tests for headless CLI export (main.run_cli_export + flag guards)."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from mypyskindose.main import run_cli_export, validate_export_flags

_RDSR = Path(__file__).resolve().parents[2] / "src" / "mypyskindose" / "example_data" / "RDSR" / "siemens_axiom_artis.dcm"
_SETTINGS = Path(__file__).resolve().parents[2] / "src" / "mypyskindose" / "settings_example.json"


@pytest.mark.skipif(not _RDSR.exists(), reason="example RDSR missing")
def test_cli_export_xlsx(tmp_path):
    out = tmp_path / "report.xlsx"
    result = run_cli_export([str(_RDSR)], str(_SETTINGS), "xlsx", export_path=out)
    assert result == out
    wb = load_workbook(out)
    assert "Overview" in wb.sheetnames
    # Events processed should be nonzero (real event table, not empty).
    ws = wb["Results"]
    values = {row[0].value: row[1].value for row in ws.iter_rows() if row[0].value}
    assert values.get("Events processed") == "21"


@pytest.mark.skipif(not _RDSR.exists(), reason="example RDSR missing")
def test_cli_export_requires_explicit_path():
    with pytest.raises(ValueError, match="--export-path is required"):
        run_cli_export([str(_RDSR)], str(_SETTINGS), "pdf")


@pytest.mark.skipif(not _RDSR.exists(), reason="example RDSR missing")
def test_cli_export_refuses_overwrite_without_force(tmp_path):
    out = tmp_path / "report.html"
    out.write_text("existing", encoding="utf-8")
    with pytest.raises(ValueError, match="export_destination_exists"):
        run_cli_export([str(_RDSR)], str(_SETTINGS), "html", export_path=out)
    run_cli_export([str(_RDSR)], str(_SETTINGS), "html", export_path=out, force=True)
    assert out.read_bytes().lower().startswith(b"<!doctype html>")


def test_cli_export_flag_conflicts():
    with pytest.raises(ValueError):
        validate_export_flags("xlsx", aggregate_only=True, input_preview_only=False, has_files=True)
    with pytest.raises(ValueError):
        validate_export_flags("xlsx", aggregate_only=False, input_preview_only=True, has_files=True)
    with pytest.raises(ValueError):
        validate_export_flags("xlsx", aggregate_only=False, input_preview_only=False, has_files=False)
    # Valid combo does not raise.
    validate_export_flags("xlsx", aggregate_only=False, input_preview_only=False, has_files=True)
    # No export format -> no-op.
    validate_export_flags(None, aggregate_only=True, input_preview_only=True, has_files=False)
