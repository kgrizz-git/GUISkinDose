"""Phase 2 tests for the XLSX report writer."""

from __future__ import annotations

import base64
import io
from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from mypyskindose import PyskindoseSettings, load_settings_example_json
from mypyskindose.export import ExportExamSource, ExportSource, collect_export_payload
from mypyskindose.export.writers.xlsx import render_xlsx_bytes

# 1x1 PNG.
_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


def _settings():
    return PyskindoseSettings(settings=load_settings_example_json(), output_format="dict")


def _out(kerma):
    n = 3
    return {
        "psd": 1.0,
        "air_kerma": float(sum(kerma)),
        "patient": {
            "patient_type": "human",
            "patient": {
                "human_phantom": "hudfrid",
                "patient_skin_cells": {"x": [0.0, 1.0, 2.0], "y": [0.0] * n, "z": [0.0] * n},
                "triangle_vertex_indices": {"i": [0], "j": [1], "k": [2]},
            },
            "orientation": "hfs",
            "offsets": {},
        },
        "dose_map": [[0, 0.5], [1, 1.0], [2, 0.3]],
        "corrections": {
            "correction_value_index": [[0, 1], [2]],
            "backscatter": [[1.2, 1.4], [1.5]],
            "inverse_square_law": [[0.9, 0.8], [1.0]],
            "medium": [1.02, 1.03],
            "table": [0.8, 0.8],
            "kerma": kerma,
        },
    }


def _fake_obj(out):
    c = out["corrections"]
    dense = np.zeros(3)
    for i, d in out["dose_map"]:
        dense[i] = d
    return SimpleNamespace(
        PSD=out["psd"], AirKerma=out["air_kerma"],
        Patient={"patient": SimpleNamespace(to_dict=lambda p=out["patient"]["patient"]: p)},
        DoseMap=dense, Hits=c["correction_value_index"], BackscatterCorrection=c["backscatter"],
        InverseSquareLawCorrection=c["inverse_square_law"], MediumCorrection=c["medium"],
        TableCorrection=c["table"], Events=SimpleNamespace(kerma=c["kerma"]),
    )


def _single_payload(with_images=False):
    src = ExportSource(
        execution_context="cli", output_dict=_out([2.0, 1.0]),
        exams=[ExportExamSource("exam1", pd.DataFrame(), None, "exam1.dcm", _settings(), (0, 0, 0))],
        file_name="exam1.dcm",
    )
    return collect_export_payload(src, with_images=with_images)


def _multi_payload():
    o1, o2 = _out([2.0, 1.0]), _out([2.0, 1.0])
    result = SimpleNamespace(
        exams=[
            SimpleNamespace(exam_id="Exam 1", source_file="a.dcm", patient_offset=[0, 0, 0], output=_fake_obj(o1), warnings=[]),
            SimpleNamespace(exam_id="Exam 2", source_file="b.dcm", patient_offset=[0, 0, 0], output=_fake_obj(o2), warnings=[]),
        ],
        aggregate_dose_map=np.array([1.0, 2.0, 0.6]), aggregate_psd=2.0, total_events=4, warnings=[],
    )
    src = ExportSource(
        execution_context="cli", multi_exam_result=result,
        exams=[
            ExportExamSource("Exam 1", pd.DataFrame(), None, "a.dcm", _settings(), (0, 0, 0)),
            ExportExamSource("Exam 2", pd.DataFrame(), None, "b.dcm", _settings(), (0, 0, 0)),
        ],
    )
    return collect_export_payload(src, with_images=False)


def test_write_xlsx_single_exam():
    wb = load_workbook(io.BytesIO(render_xlsx_bytes(_single_payload())))
    for sheet in ("Overview", "Results", "Settings", "Corrections", "Warnings", "Images"):
        assert sheet in wb.sheetnames


def test_write_xlsx_multi_exam_summary_columns():
    wb = load_workbook(io.BytesIO(render_xlsx_bytes(_multi_payload())))
    ws = wb["Results"]
    header = [c.value for c in ws[1]]
    assert "Exam 1" in header
    assert "Exam 2" in header
    assert "Cumulative" in header


def test_write_xlsx_images_embedded():
    payload = _single_payload()
    payload.images = [SimpleNamespace(label="Dorsal", view="dorsal", exam_id=None, png_bytes=_PNG, error_message=None)]
    wb = load_workbook(io.BytesIO(render_xlsx_bytes(payload)))
    assert wb["Images"]._images  # anchored drawings present


def test_write_xlsx_missing_images():
    payload = _single_payload()
    payload.images = [
        SimpleNamespace(label="Dorsal", view="dorsal", exam_id=None, png_bytes=None,
                        error_message="Image unavailable (kaleido/export error)")
    ]
    data = render_xlsx_bytes(payload)  # must not raise
    wb = load_workbook(io.BytesIO(data))
    assert not wb["Images"]._images
