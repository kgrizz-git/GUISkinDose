"""Tests for spreadsheet formula-injection neutralization (CWE-1236)."""

from __future__ import annotations

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from mypyskindose.spreadsheet_safety import neutralize_dataframe, neutralize_spreadsheet_value


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ('=HYPERLINK("https://attacker.example","open")', '\'=HYPERLINK("https://attacker.example","open")'),
        ("+cmd|' /C calc'!A0", "'+cmd|' /C calc'!A0"),
        ("-2+3+cmd|' /C calc'!A0", "'-2+3+cmd|' /C calc'!A0"),
        ("@SUM(1+1)", "'@SUM(1+1)"),
        ("\tbeacon", "'\tbeacon"),
        ("\rbeacon", "'\rbeacon"),
        ("normal text", "normal text"),
        ("", ""),
    ],
)
def test_neutralize_spreadsheet_value_prefixes(raw: str, expected: str) -> None:
    assert neutralize_spreadsheet_value(raw) == expected


def test_neutralize_spreadsheet_value_preserves_non_strings() -> None:
    assert neutralize_spreadsheet_value(42) == 42
    assert neutralize_spreadsheet_value(None) is None


def test_neutralize_dataframe_only_touches_object_columns() -> None:
    df = pd.DataFrame(
        {
            "VendorNote": ['=HYPERLINK("x","y")', "ok"],
            "kVp": [80.0, 90.0],
        }
    )
    safe = neutralize_dataframe(df)
    vendor_note = safe.loc[0, "VendorNote"]
    assert isinstance(vendor_note, str)
    assert vendor_note.startswith("'")
    assert safe.loc[1, "VendorNote"] == "ok"
    assert safe["kVp"].tolist() == [80.0, 90.0]
    raw_vendor_note = df.loc[0, "VendorNote"]
    assert isinstance(raw_vendor_note, str)
    assert raw_vendor_note.startswith("=")


def test_neutralize_dataframe_prefixes_dangerous_column_names() -> None:
    payload = '=HYPERLINK("https://attacker.example/beacon","open")'
    df = pd.DataFrame([{payload: 1, "safe": 2}])
    safe = neutralize_dataframe(df)
    assert list(safe.columns) == ["'" + payload, "safe"]
    assert list(df.columns) == [payload, "safe"]


def test_neutralize_dataframe_prefixes_dangerous_index_labels() -> None:
    df = pd.DataFrame({"kVp": [80.0]}, index=pd.Index(['=CMD|"/C calc"!A0'], name="@SUM(1)"))
    safe = neutralize_dataframe(df)
    assert safe.index.name == "'@SUM(1)"
    assert safe.index.tolist() == ["'=CMD|\"/C calc\"!A0"]


def test_neutralize_dataframe_preserves_and_neutralizes_column_axis_name() -> None:
    df = pd.DataFrame([[1, 2]], columns=pd.Index(["a", "=CMD"], name="@axis"))
    safe = neutralize_dataframe(df)
    assert list(safe.columns) == ["a", "'=CMD"]
    assert safe.columns.name == "'@axis"


def test_neutralize_dataframe_preserves_multiindex_column_names() -> None:
    columns = pd.MultiIndex.from_tuples(
        [("grp", "=A1"), ("grp", "safe")],
        names=["@level0", "level1"],
    )
    df = pd.DataFrame([[1, 2]], columns=columns)
    safe = neutralize_dataframe(df)
    assert safe.columns.names == ["'@level0", "level1"]
    assert list(safe.columns) == [("grp", "'=A1"), ("grp", "safe")]


def test_neutralize_dataframe_preserves_multiindex_row_names() -> None:
    index = pd.MultiIndex.from_tuples(
        [("grp", "=A1"), ("grp", "safe")],
        names=["@level0", "level1"],
    )
    df = pd.DataFrame({"kVp": [80.0, 90.0]}, index=index)
    safe = neutralize_dataframe(df)
    assert safe.index.names == ["'@level0", "level1"]
    assert list(safe.index) == [("grp", "'=A1"), ("grp", "safe")]
    assert list(df.index) == [("grp", "=A1"), ("grp", "safe")]
    assert df.index.names == ["@level0", "level1"]


def test_neutralized_xlsx_export_is_not_formula_cell() -> None:
    payload = '=HYPERLINK("https://attacker.example/beacon","open")'
    safe_df = neutralize_dataframe(pd.DataFrame([{"VendorNote": payload}]))
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False)
    buf.seek(0)
    ws = load_workbook(buf)["Sheet1"]
    cell = ws["A2"]
    assert cell.data_type == "s"
    assert cell.value.startswith("'")


def test_neutralized_xlsx_header_is_not_formula_cell() -> None:
    payload = '=HYPERLINK("https://attacker.example/beacon","open")'
    safe_df = neutralize_dataframe(pd.DataFrame([{payload: "ok"}]))
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        safe_df.to_excel(writer, index=False)
    buf.seek(0)
    ws = load_workbook(buf)["Sheet1"]
    header = ws["A1"]
    assert header.data_type == "s"
    assert isinstance(header.value, str)
    assert header.value.startswith("'")
